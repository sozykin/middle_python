# Подключаем библиотеку openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font

# Создаем книгу Excel
wb = Workbook()

# Получаем активный лист
ws = wb.active

# Вставляем заголовки
ws.append(['Название', 'Описание'])
# Выделяем заголовки жирный шрифтом
ft = Font(bold=True)
ws['A1'].font = ft
ws['B1'].font = ft

# Добавляем строки таблицы
ws.append(['HTTP', 'Протокол передачи гипертекста'])
ws.append(['SMTP', 'Простой протокол передачи почты'])

# Задаем ширину столбцов
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 50

# Сохраняем книгу Excel
wb.save('demo.xlsx')
